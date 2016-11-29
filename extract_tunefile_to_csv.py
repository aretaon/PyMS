# -*- coding: utf-8 -*-
"""
Created on Tue Nov 29 09:21:26 2016

@author: aretaon
"""

import argparse
import os
import codecs
import pprint

pp = pprint.PrettyPrinter(indent=4)

import csv

from openpyxl import load_workbook

import time

thisdate = time.strftime("%Y%m%d")

DataDirs = ["/home/aretaon/AGW/RawData/Synapt",
            "/home/aretaon/AGW/RawData/Ultima"]

parser = argparse.ArgumentParser()

parser.add_argument("-f", "--files", help="Path to a file containing the basenames\
                                           of the Waters rawfiles to parse")

parser.add_argument("-i", "--instrumentlists", help="A comma separated list\
                                                     of xlsx-files\
                                                     to look\
                                                     for further description",
                                               type=str)

args = parser.parse_args()

if args.files != None:

    with open(args.files, 'r') as l:
        FilesToLookFor = l.readlines()
    
    FilesToLookFor = [f.strip() for f in FilesToLookFor]
    
else:
    print('Collecting all tune-data!')


rawfiles = []

for d in DataDirs:
    for root, dirs, files in os.walk(d):
        for f in dirs:
            if args.files != None:
                if f.endswith('.raw')\
                   and f[:-4] in FilesToLookFor:
                       print("Found {:s}!".format(f))
                       rawfiles.append(os.path.join(root, f))
            else:
                if f.endswith('.raw'):
                       rawfiles.append(os.path.join(root, f))

base_rawfiles = [os.path.split(f)[1][:-4] for f in rawfiles]

if args.files != None:
    outname = '{:s}_{:s}_tune.csv'.format(thisdate, '_'.join(base_rawfiles))
else:
    outname = '{:s}_complete_tune.csv'.format(thisdate)

outhandler = codecs.open(outname,
                         'w',
                         encoding='utf-8')


instrumentlists = list(args.instrumentlists.split(','))

i_list = []

for li in instrumentlists:
    wb = load_workbook(filename=li, read_only=True)
    # The lists are called Data on both instruments
    i_list.append(wb['Data'])
    
DictList = []

for r in rawfiles:

    ParamDict = {}
    
    ParamDict['Rawfile'] = os.path.split(r)[1]
    
    # encoding has to be set to latin1 as MassLynx does not use UTF-8
    try:
        inhandler = codecs.open(os.path.join(r,'_extern.inf'),
                            'r',
                            encoding='latin1')
        extern = inhandler.readlines()
        inhandler.close()
    
        for line in extern:
            if '\t' in line:
                key = line.strip().split('\t')[0]
                value = line.strip().split('\t')[-1]
                ParamDict[key] = value

        # extract the corresponding data from the excel lists

        for sheet in i_list:
            i = 0
            header = []
            for j, row in enumerate(sheet.iter_rows()):
                # get the header
                if j == 0:
                    for cell in row:
                        if cell.value != None:
                            header.append(cell.value)
                
                RightRow = False
                for idx, cell in enumerate(row):
                    if cell.value == os.path.split(r)[1][:-4]:
                        RightRow = True
                    if RightRow and idx < header.index('Messzeit [min]'):
                        ParamDict[header[idx]] = cell.value
                            
        DictList.append(ParamDict)
        
    except FileNotFoundError as e:
        print(e)

fieldnames = set()

for d in DictList:
    # use the union parameter to add the dict to the fieldnames set
    fieldnames |= set(d.keys())

# generate an order to print
fieldnames = sorted(list(fieldnames))
rawfile_idx = fieldnames.index('Rawfile')
# move the rawfile ID at the first position
fieldnames[0], fieldnames[rawfile_idx] = fieldnames[rawfile_idx], fieldnames[0]

#pp.pprint(DictList)

writer = csv.DictWriter(outhandler,
                        fieldnames=fieldnames,
                        dialect='excel-tab')
writer.writeheader()
for d in DictList:
    writer.writerow(d)
    
outhandler.close()