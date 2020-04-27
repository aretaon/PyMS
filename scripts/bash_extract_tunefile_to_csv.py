# -*- coding: utf-8 -*-
"""
Created on Tue Nov 29 09:21:26 2016

@author: aretaon
"""

import argparse
import os
import codecs
import pprint

pp = pprint.PrettyPrinter(indent=4)

import csv

from openpyxl import load_workbook

import time

thisdate = time.strftime("%Y%m%d")

DataDirs = ["/home/aretaon/AGW/RawData/Synapt",
            "/home/aretaon/AGW/RawData/Ultima"]

parser = argparse.ArgumentParser()

parser.add_argument("-f", "--files", help="Path to a file containing the basenames\
                                           of the Waters rawfiles to parse")

parser.add_argument("-i", "--instrumentlists", help="A comma separated list\
                                                     of xlsx-files\
                                                     to look\
                                                     for further description",
                                               type=str)

args = parser.parse_args()

if args.files != None:

    with open(args.files, 'r') as l:
        FilesToLookFor = l.readlines()
    
    FilesToLookFor = [f.strip() for f in FilesToLookFor]
    
else:
    print('Collecting all tune-data!')


rawfiles = []

for d in DataDirs:
    for root, dirs, files in os.walk(d):
        for f in dirs:
            if args.files != None:
                if f.endswith('.raw')\
                   and f[:-4] in FilesToLookFor:
                       print("Found {:s}!".format(f))
                       rawfiles.append(os.path.join(root, f))
            else:
                if f.endswith('.raw'):
                       rawfiles.append(os.path.join(root, f))

base_rawfiles = [os.path.split(f)[1][:-4] for f in rawfiles]

if args.files != None and len(rawfiles) < 4:
    outname = '{:s}_{:s}_tune.tsv'.format(thisdate, '_'.join(base_rawfiles))
elif args.files != None:
    outname = '{:s}_multiple_tune.tsv'.format(thisdate)
else:
    outname = '{:s}_complete_tune.tsv'.format(thisdate)

outhandler = codecs.open(outname,
                         'w',
                         encoding='utf-8')


instrumentlists = list(args.instrumentlists.split(','))

i_list = []

for li in instrumentlists:
    wb = load_workbook(filename=li, read_only=True)
    # The lists are called Data on both instruments
    i_list.append(wb['Data'])
    
DictList = []

for r in rawfiles:

    ParamDict = {}
    
    ParamDict['Rawfile'] = os.path.split(r)[1]
    
    # encoding has to be set to latin1 as MassLynx does not use UTF-8
    try:
        inhandler = codecs.open(os.path.join(r,'_extern.inf'),
                            'r',
                            encoding='latin1')
        extern = inhandler.readlines()
        inhandler.close()
    
        for line in extern:
            if '\t' in line:
                key = line.strip().split('\t')[0]
                value = line.strip().split('\t')[-1]
                ParamDict[key] = value

        # extract the corresponding data from the excel lists
        EndReached = False

        for sheet in i_list:
            if EndReached == True:
                #print('Reached end of {}'.format(sheet))
                break
            
            i = 0
            header = []
            # avoid long searching in empty datasheets: End the search
            # after three consecutive empty rows
            EmptyRows = 0
            for j, row in enumerate(sheet.iter_rows()):
                if EndReached == True:
                    break
                # get the header
                if j == 0:
                    for cell in row:
                        if cell.value != None:
                            header.append(cell.value)
                
                RightRow = False
                for idx, cell in enumerate(row):
                    if idx == 0 and cell.value == 'none':
                        #print('now found an empty cell at row {} cell {}'.format(j, idx))
                        EmptyRows += 1
                        #print('Empty rows is {}'.format(EmptyRows))
                        if EmptyRows > 4:
                            # Restrict the search horizontally (by rows)
                            EndReached = True
                        break
                    if cell.value != 'none':
                        EmptyRows = 0
                    if cell.value == os.path.split(r)[1][:-4]:
                        RightRow = True
                    if RightRow and idx < header.index('Messzeit [min]'):
                        ParamDict[header[idx]] = cell.value
                    elif idx > header.index('Messzeit [min]'):
                        # restrict the search vertically (by column)
                        break
        DictList.append(ParamDict)
        
    except FileNotFoundError as e:
        print(e)

fieldnames = set()

for d in DictList:
    # use the union parameter to add the dict to the fieldnames set
    fieldnames |= set(d.keys())

# generate an order to print
fieldnames = sorted(list(fieldnames))

def MoveForward(name, position, data):
    """
    Move a list entry at a specified position
    
    Requires:
    - name: Name of entry header
    - position: Desired final position
    - data: Pandas dataframe
    
    Returns:
    - ordered dataframe
    """
    
    rawfile_idx = data.index(name)
    # move the rawfile ID at the indicated position
    data[position], data[rawfile_idx] = \
    data[rawfile_idx], data[position]

    return data
    
StandardFirstColumns = ['Rawfile',
                        'Probenname',
                        'Puffer',
                        'Kommentar',
                        'Backing',
                        'Capillary',
                        'Capillary (kV)',
                        'Cell pressure [e^-3 mbar]',
                        'Collision Energy',
                        'Transfer Collision Energy',
                        'Trap Collision Energy']

FirstColumns = []

for name in StandardFirstColumns:
    if name in fieldnames:
        FirstColumns.append(name)

for name, position in zip(FirstColumns, range(len(FirstColumns))):
    # names differ depending on the MS-Instrument used
    fieldnames = MoveForward(name, position, fieldnames)

#pp.pprint(DictList)

writer = csv.DictWriter(outhandler,
                        fieldnames=fieldnames,
                        dialect='excel-tab')
writer.writeheader()
for d in DictList:
    writer.writerow(d)
    
outhandler.close()