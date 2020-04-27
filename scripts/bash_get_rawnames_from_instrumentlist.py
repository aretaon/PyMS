# -*- coding: utf-8 -*-
"""
Created on Tue Dec  6 14:20:48 2016

@author: aretaon
"""

import argparse

from openpyxl import load_workbook

import time

import re

import os

thisdate = time.strftime("%Y%m%d")

DataDirs = ["/home/aretaon/AGW/RawData/Synapt",
            "/home/aretaon/AGW/RawData/Ultima"]

parser = argparse.ArgumentParser()

parser.add_argument("projects", help="Comma separated list\
                                      of projects to extract rawfile\
                                      names from",
                                type=str)

parser.add_argument("instrumentlists", help="A comma separated list\
                                                     of xlsx-files\
                                                     to look\
                                                     for further description",
                                               type=str)

# generate parser object -> extract arguments from input
args = parser.parse_args()

# Read arguments from input
print('===Instrumentslist===\n{}'.format('\n'.join(args.instrumentlists.split(','))))
instrumentlists = list(args.instrumentlists.split(','))
print('===Projects List===\n{}'.format('\n'.join(args.projects.split(','))))
Projects = list(args.projects.split(','))

i_list = []

for li in instrumentlists:
    wb = load_workbook(filename=li, read_only=True)
    # The lists are called Data on both instruments
    i_list.append(wb['Data'])
    
EndReached = False
Rawfiles = []
for sheet in i_list:
    if EndReached == True:
        #print('Reached end of {}'.format(sheet))
        break
    
    i = 0
    header = []
    # avoid long searching in empty datasheets: End the search
    # after three consecutive empty rows
    EmptyRows = 0
    for j, row in enumerate(sheet.iter_rows()):
        if EndReached == True:
            break
        # get the header
        if j == 0:
            for cell in row:
                if cell.value != None:
                    header.append(cell.value.strip())
                # avoid collecting the whole header for only a subset of columns
                if cell.value.strip() == 'True Measurement':
                    break
        
        SaveCells = []
        SaveThis = False
        for idx, cell in enumerate(row):
            if idx == 0 and cell.value == 'none':
                #print('now found an empty cell at row {} cell {}'.format(j, idx))
                EmptyRows += 1
                #print('Empty rows is {}'.format(EmptyRows))
                if EmptyRows > 4:
                    # Restrict the search horizontally (by rows)
                    EndReached = True
                break
            if cell.value != 'none':
                EmptyRows = 0
            # collect all cell values including the rawfile name
            if idx < header.index('Projektnummer'):
                SaveCells.append(cell.value)
            # check if the row is attributed to the searched project
            if idx == header.index('Projektnummer'):
                for p in Projects:
                    if p in str(cell.value):
                        #print('Matched {} to {}'.format(p, str(cell.value)))
                        SaveThis = True
            # only save if the Measurement is True
            if idx == header.index('True Measurement') and cell.value == True and SaveThis:
                Rawfiles.append(SaveCells[0])
            if idx > header.index('True Measurement'):
                # restrict the search vertically (by column)
                break
            
if len(Projects) == 1:
    outname = Projects[0]
if len(Projects) < 4:
    outname = '_'.join(Projects)
else:
    outname = '_'.join(Projects[0], 'et_al')

# remove not printable characters
outname = re.sub(r'(\W)','_', outname) + '_rawfiles.txt'

with open(outname, 'w') as f:
    for r in Rawfiles:
        f.write('{}\n'.format(r))

print('\nWrote {} rawfiles for the project{} {} to {}'.format(len(Rawfiles),
                                                            's' if len(Projects) > 1 else '',
                                                            ', '.join(Projects),
                                                            os.path.join(os.getcwd(), outname)))